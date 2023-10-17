import pandas as pd
from pulp import *
import openpyxl

# Read in the CSV data from that was exported from DraftKings
drivers = pd.read_csv(r"C:\Users\dkens\OneDrive\Desktop\DKSalaries.csv", usecols=['Name', 'Roster_Position', 'AvgPointsPerGame', 'Salary'])

# Clean up driver names by removing leading/trailing whitespaces
drivers['Name'] = drivers['Name'].str.strip()

# Create an Excel workbook and active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Group the drivers by their name, roster position, average points per race, and salary
availables = drivers.groupby(["Name", "Roster_Position", "AvgPointsPerGame", "Salary"]).agg('count')
availables = availables.reset_index()

# Create dictionaries to store salary and point information for each roster position
salaries = {}
points = {}

# Populate the salary and point dictionaries for each roster position
for pos in availables.Roster_Position.unique():
    available_pos = availables[availables.Roster_Position == pos]
    salary = list(available_pos[['Name', 'Salary']].set_index("Name").to_dict().values())[0]
    point = list(available_pos[['Name', 'AvgPointsPerGame']].set_index("Name").to_dict().values())[0]

    salaries[pos] = salary
    points[pos] = point

# Number of Roster_Positions available in a DFS Lineup
pos_num_available = {
    "CPT": 1,
    "D": 4,
    "CNSTR": 1
}

# Define the Salary Cap for the contest
salary_cap = 50000

# Budget constraint for the CPT
CPT_budget = 24000  # You can adjust this value based on the desired budget for the CPT, this is the max for the current contest 

# The amount of lineups run in a simulation
for lineup in range(1, 21):
    # Establish variables for our pulp problem
    _vars = {k: LpVariable.dict(k, v, cat="Binary") for k, v in points.items()}

    prob = LpProblem("Fantasy", LpMaximize)
    rewards = []
    costs = []
    Roster_Position_constraints = []
    # Initialize a set to keep track of drivers selected as CPT
    selected_CPTs = set()

    # Loop through each roster position and add constraints
    for k, v in _vars.items():
        costs += lpSum([salaries[k][i] * _vars[k][i] for i in v])
        rewards += lpSum([points[k][i] * _vars[k][i] for i in v])

        if k == "CPT":
            # Ensure exactly 1 CPT is picked and impose budget constraint
            prob += lpSum([_vars[k][i] for i in v]) == pos_num_available[k]
            prob += lpSum([salaries[k][i] * _vars[k][i] for i in v]) <= CPT_budget

            # Add a constraint to prevent a driver selected as CPT from being chosen as a Driver
            for driver in v:
                print(driver)  # This print statement is to identify the specific driver causing the KeyError 
                prob += _vars["D"][driver] + _vars["CPT"][driver] <= 1
                if _vars["CPT"][driver].value() == 1:  # Add the selected CPT to the set - Thank you StackOverflow
                    selected_CPTs.add(driver)

        else:
            prob += lpSum([_vars[k][i] for i in v]) == pos_num_available[k]

    # Add a constraint to ensure that the CPT selected in the lineup are not picked again
    for driver in selected_CPTs:
        prob += _vars["CPT"][driver] <= 1

   # Add up the rewards (points) for all selected players in the lineup
    prob += lpSum(rewards)

    # Ensure that the total cost of the lineup does not exceed the salary cap
    prob += lpSum(costs) <= salary_cap

    # Additional constraint to prevent duplicate lineups 
    # If it's not the first lineup, make sure that the total score of the current lineup is less than the previous one by a small margin (0.001)
    # This helps to find slightly different lineups in each iteration
    if not lineup == 1:
        prob += (lpSum(rewards) <= total_score - 0.001)

    # Solve the optimization problem to find the best lineup
    prob.solve()

    # Get the score (objective value) of the optimized lineup
    score = str(prob.objective)

    # Get the names of the constraints used in the optimization problem (for debugging or reference purposes)
    constants = [str(const) for const in prob.constraints.values()]

    # Initialize the column number for writing the selected lineup to the Excel worksheet
    colnum = 1

    # Write the selected lineup to the Excel worksheet
    for v in prob.variables():
        score = score.replace(v.name, str(v.varValue))
        if v.varValue != 0:
            ws.cell(row=lineup, column=colnum).value = v.name
            colnum += 1
    total_score = eval(score)
    ws.cell(row=lineup, column=colnum).value = total_score
    print(lineup, total_score)

# Save the Excel workbook
wb.save(r"C:\Users\dkens\OneDrive\Desktop\Lineups.xls")
